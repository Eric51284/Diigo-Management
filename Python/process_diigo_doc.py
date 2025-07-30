import pandas as pd
from docx import Document
import re
from difflib import SequenceMatcher
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment


def extract_all_articles_from_docx(docx_path):
    """
    More comprehensive extraction that catches all article entries
    """
    doc = Document(docx_path)
    outline_data = []
    current_hierarchy = ["", "", "", ""]  # Main, Sub1, Sub2, Sub3

    print("Analyzing document structure...")

    for para_idx, paragraph in enumerate(doc.paragraphs):
        text = paragraph.text.strip()

        if not text:
            continue

        print(f"Para {para_idx}: Level ? - '{text[:80]}...'")

        # Get the actual outline level from Word
        level = get_word_outline_level(paragraph)

        # Alternative: analyze the text pattern
        if level is None:
            level = analyze_text_pattern(text)

        print(f"  -> Determined level: {level}")

        # Check if this looks like an article entry (has date pattern)
        date_match = re.search(r"(\d{4}-\d{2}-\d{2}):\s*(.*)", text)

        if date_match:
            # This is an article entry
            date = date_match.group(1)
            title = date_match.group(2).strip()

            # Extract hyperlink if present
            hyperlink_url = extract_hyperlink_from_paragraph(paragraph)

            # Clean up hierarchy for this article
            current_clean_hierarchy = [h for h in current_hierarchy if h]

            outline_data.append(
                {
                    "main_category": (
                        current_hierarchy[0]
                        if current_hierarchy[0]
                        else "Uncategorized"
                    ),
                    "subcategory_1": (
                        current_hierarchy[1] if len(current_hierarchy) > 1 else ""
                    ),
                    "subcategory_2": (
                        current_hierarchy[2] if len(current_hierarchy) > 2 else ""
                    ),
                    "subcategory_3": (
                        current_hierarchy[3] if len(current_hierarchy) > 3 else ""
                    ),
                    "date": date,
                    "title": title,
                    "doc_hyperlink": hyperlink_url,
                    "full_hierarchy": " > ".join(current_clean_hierarchy),
                    "raw_text": text,
                }
            )
            print(f"  -> ARTICLE FOUND: {title[:50]}...")

        else:
            # This is a category/subcategory header
            clean_text = text.replace("-", "").strip()

            # Update hierarchy based on level
            if level is not None:
                # Reset deeper levels
                for i in range(level, len(current_hierarchy)):
                    current_hierarchy[i] = ""

                # Set current level
                if level < len(current_hierarchy):
                    current_hierarchy[level] = clean_text

                print(f"  -> CATEGORY: Level {level} = '{clean_text}'")
                print(f"  -> Current hierarchy: {current_hierarchy}")

    print(f"\nTotal articles found: {len(outline_data)}")
    return outline_data


def get_word_outline_level(paragraph):
    """
    Try to get the actual outline level from Word formatting
    """
    try:
        # Check for outline level in paragraph properties
        if paragraph._element.pPr is not None:
            outline_lvl = paragraph._element.pPr.find(
                ".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}outlineLvl"
            )
            if outline_lvl is not None:
                return int(
                    outline_lvl.get(
                        "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val",
                        0,
                    )
                )

            # Check for numbering properties
            numPr = paragraph._element.pPr.find(
                ".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}numPr"
            )
            if numPr is not None:
                ilvl = numPr.find(
                    ".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}ilvl"
                )
                if ilvl is not None:
                    return int(
                        ilvl.get(
                            "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val",
                            0,
                        )
                    )

        # Check style-based outline level
        if paragraph.style.name.startswith("Heading"):
            return int(paragraph.style.name.replace("Heading ", "")) - 1

    except Exception as e:
        print(f"Error getting outline level: {e}")

    return None


def analyze_text_pattern(text):
    """
    Analyze text pattern to determine hierarchy level
    """
    # Count leading spaces and dashes
    leading_spaces = len(text) - len(text.lstrip(" "))

    # Different patterns for different levels
    if re.match(r"^[A-Z][a-z].*", text.strip()) and leading_spaces == 0:
        # Looks like main category (starts with capital, no indentation)
        return 0
    elif text.strip().startswith("-") and leading_spaces <= 4:
        # First level bullet
        return 1
    elif text.strip().startswith("-") and leading_spaces > 4 and leading_spaces <= 8:
        # Second level bullet
        return 2
    elif text.strip().startswith("-") and leading_spaces > 8:
        # Third level bullet
        return 3
    elif re.match(r"\d{4}-\d{2}-\d{2}:", text):
        # Article entry - don't change hierarchy
        return None
    else:
        # Try to guess based on content
        if any(word in text.lower() for word in ["generative ai", "main", "section"]):
            return 0
        else:
            return 1


def extract_hyperlink_from_paragraph(paragraph):
    """
    Extract hyperlink URL from paragraph
    """
    try:
        # Look for hyperlinks in runs
        for run in paragraph.runs:
            if run._element.rPr is not None:
                for elem in run._element.iter():
                    if "hyperlink" in str(elem.tag).lower():
                        # Try to get the relationship ID
                        r_id = elem.get(
                            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                        )
                        if (
                            r_id
                            and hasattr(paragraph.part, "rels")
                            and r_id in paragraph.part.rels
                        ):
                            return paragraph.part.rels[r_id].target_ref

        # Alternative method: look for hyperlink elements directly
        for elem in paragraph._element.iter():
            if "hyperlink" in str(elem.tag).lower():
                r_id = elem.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                if (
                    r_id
                    and hasattr(paragraph.part, "rels")
                    and r_id in paragraph.part.rels
                ):
                    return paragraph.part.rels[r_id].target_ref

    except Exception as e:
        print(f"Error extracting hyperlink: {e}")

    return None


def similarity_score(a, b):
    """Calculate similarity between two strings"""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def match_urls_from_csv(outline_data, csv_path):
    """Match outline items with URLs from CSV file"""
    print(f"Loading CSV data from {csv_path}...")
    df = pd.read_csv(csv_path)
    print(f"CSV contains {len(df)} entries")

    matched_count = 0

    for i, item in enumerate(outline_data):
        if i % 50 == 0:
            print(f"Processing item {i+1}/{len(outline_data)}")

        best_match = None
        best_score = 0
        best_url = None
        best_csv_title = None

        outline_title = item["title"]

        # Search through CSV for best title match
        for _, row in df.iterrows():
            csv_title = str(row["title"])

            # Skip if CSV title is NaN or empty
            if pd.isna(csv_title) or not csv_title.strip():
                continue

            # Calculate similarity score
            score = similarity_score(csv_title, outline_title)

            # Bonus for date match
            date_bonus = 0
            if "created_at" in row and pd.notna(row["created_at"]):
                csv_date = str(row["created_at"])[:10]  # Extract YYYY-MM-DD
                if csv_date == item["date"]:
                    date_bonus = 0.3

            total_score = score + date_bonus

            if total_score > best_score:
                best_score = total_score
                best_match = csv_title
                best_url = row["url"]
                best_csv_title = csv_title

        # Add match results
        item["matched_title"] = best_match
        item["match_score"] = best_score
        item["csv_url"] = best_url if best_score > 0.5 else None  # Lower threshold
        item["final_url"] = (
            item["csv_url"] if item["csv_url"] else item["doc_hyperlink"]
        )

        if item["csv_url"]:
            matched_count += 1

    print(f"Matched {matched_count} items with CSV data")
    return outline_data


def create_excel_file(outline_data, output_path):
    """Create Excel file with outline structure and hyperlinks"""
    # Create DataFrame
    df = pd.DataFrame(outline_data)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Research Outline"

    # Headers
    headers = [
        "Main Category",
        "Subcategory 1",
        "Subcategory 2",
        "Subcategory 3",
        "Date",
        "Title",
        "URL",
        "Match Score",
        "Source",
        "Full Hierarchy",
    ]

    # Add headers with formatting
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

    # Add data
    for row_idx, item in enumerate(outline_data, 2):
        ws.cell(row=row_idx, column=1, value=item["main_category"])
        ws.cell(row=row_idx, column=2, value=item["subcategory_1"])
        ws.cell(row=row_idx, column=3, value=item["subcategory_2"])
        ws.cell(row=row_idx, column=4, value=item["subcategory_3"])
        ws.cell(row=row_idx, column=5, value=item["date"])

        # Create hyperlinked title
        title_cell = ws.cell(row=row_idx, column=6, value=item["title"])
        if item["final_url"]:
            title_cell.hyperlink = item["final_url"]
            title_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=row_idx, column=7, value=item["final_url"] or "")
        ws.cell(row=row_idx, column=8, value=round(item.get("match_score", 0), 3))

        # Source
        source = (
            "CSV Match"
            if item["csv_url"]
            else "Doc Link" if item["doc_hyperlink"] else "No Link"
        )
        ws.cell(row=row_idx, column=9, value=source)
        ws.cell(row=row_idx, column=10, value=item["full_hierarchy"])

    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")


def main(docx_path, csv_path, output_path):
    """Main function"""

    print("Step 1: Extracting ALL articles from Word document...")
    outline_data = extract_all_articles_from_docx(docx_path)

    print("Step 2: Matching with CSV data...")
    matched_data = match_urls_from_csv(outline_data, csv_path)

    print("Step 3: Creating Excel file...")
    create_excel_file(matched_data, output_path)

    # Summary
    total = len(matched_data)
    csv_matches = sum(1 for item in matched_data if item["csv_url"])
    doc_links = sum(
        1 for item in matched_data if item["doc_hyperlink"] and not item["csv_url"]
    )
    no_links = sum(1 for item in matched_data if not item["final_url"])

    print(f"\n=== SUMMARY ===")
    print(f"Total articles found: {total}")
    print(f"CSV matches: {csv_matches}")
    print(f"Document links: {doc_links}")
    print(f"No links: {no_links}")



if __name__ == "__main__":
    main()
